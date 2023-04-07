# Прочитать сохранённый csv-файл из предыдущего задания и сохранить
# данные в excel-файл, кроме возраста – столбец с этими данными не нужен.

import openpyxl
import csv

with open('data.csv', 'r') as csv_file:
    csv_reader = csv.reader(csv_file)

    wb = openpyxl.Workbook()
    ws = wb.active
    rows = list(csv_reader)
    num_rows = sum(1 for row in csv_reader)

    for r, row in enumerate(rows, start=1):
        ws.cell(row=1, column=1, value='')
        ws.cell(row=1, column=r, value='Person ' + str(r - 1))
        for c, val in enumerate(row, start=2):
            if c == 4:
                continue
            ws.cell(row=c if c < 4 else c - 1, column=r, value=val)

    wb.save('dataexel.xlsx')
    wb.close()
