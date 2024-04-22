# Використовуючи дані з минулого прикладу, спочатку записати їх у файл
# Excel-формату, а потім зчитати зі збереженого файлу

import openpyxl

name_of_fields = ["Ім'я", "Клас", "Вік"]
field_01 = ["Євген", 3, 10]
field_02 = ["Сашко", 5, 12]
field_03 = ["Михайло", 11, 18]

wb = openpyxl.Workbook()
print(wb.sheetnames)

wb.create_sheet(title='Перший аркуш', index=0)
print(wb.sheetnames)
sheet = wb['Sheet']
print(sheet)

wb.remove(sheet)
print(wb.sheetnames)

sheet = wb['Перший аркуш']
print(sheet)

for row_index, row in enumerate((name_of_fields, field_01, field_02, field_03)):
    for col_index, value in enumerate(row):
        cell = sheet.cell(row=row_index+1, column=col_index+1)
        cell.value = value

wb.save('task_04.xlsx')
