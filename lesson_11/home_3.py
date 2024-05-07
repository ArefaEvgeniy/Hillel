# Прочитати збережений csv-файл із попереднього завдання та зберегти дані
# в excel-файл. Крім віку - стовпець з цими даними не потрібний.
# *Додаткове завдання не обов'язкове до виконання:
# розгорнути таблицю на дев'яносто градусів (стовпці стають рядками,
# а рядки стовпцями). До завдання прикріплений приклад як має виглядати
# змісту підсумкового файлу.
# Прочитати збережений csv-файл із попереднього завдання та зберегти дані
# в excel-файл. Крім віку - стовпець з цими даними не потрібний.

import csv
import openpyxl

input_data = []
names = None

# відкриття файлу для читання
with open("file_csv.csv", encoding='utf-8') as r_file:
    file_reader = csv.reader(r_file, delimiter=",")
    count = 0
    for row in file_reader:
        if count == 0:
            names = row  # вичитування найменувань стовпців (перший рядок у файлі)
        else:
            input_data.append(row)  # зчитування даних рядково
        count += 1

# знаходження індексу стовпця 'age'
age_index = [index for index, item in enumerate(names) if item == 'age']
age_index = age_index[0] if age_index else None

# якщо цей стовпець є, то видалити його значення зі списку найменувань і зі списку даних
if age_index is not None:
    names.pop(age_index)
    for item_list in input_data:
        item_list.pop(age_index)

wb = openpyxl.Workbook()  # відкриття книги excel
sheet = wb.active  # вибір активного листа

# Запис верхнього рядка person 1, person 2 ... починаючи з другого стовпця
# Кількість записів person дорівнює кількості списків всередині input_data
for item in range(1, len(input_data) + 1):
    cell = sheet.cell(row=1, column=item + 1)
    cell.value = f'person {item}'

# цикл перебору рядків
for name_index, name in enumerate(names):
    cell = sheet.cell(row=name_index+2, column=1)
    cell.value = name
    # цикл перебору стовпців
    for item in range(1, len(input_data) + 1):
        cell = sheet.cell(row=name_index+2, column=item + 1)
        cell.value = input_data[item - 1][name_index]

wb.save('file_xlsx.xlsx')
