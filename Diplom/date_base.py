import openpyxl
import json
import time

from person import Person


class ColumnException(Exception):
    pass


class DB(object):

    name_fields = ('Name', 'Surname', 'Lastname', 'Birthday', 'Death', 'Sex')
    extensions = {3: ".xlsx", 4: ".xlsx", 5: ".json"}

    def __init__(self):
        self.data = []
        self.choices = {3: self.__load, 4: self.__save, 5: self.__export}

    def __load(self, f_name):
        if f_name:
            try:
                wb = openpyxl.load_workbook(f_name)
                sheet = wb.active
                rows = sheet.max_row
                cols = sheet.max_column
                if cols != 6:
                    raise ColumnException
                self.data = []
                for i in range(2, rows + 1):
                    name = sheet.cell(row=i, column=1).value
                    s_name = sheet.cell(row=i, column=2).value
                    s_name = s_name if s_name is not None else ''
                    l_name = sheet.cell(row=i, column=3).value
                    l_name = l_name if l_name is not None else ''
                    birthday = sheet.cell(row=i, column=4).value
                    death = sheet.cell(row=i, column=5).value
                    death = death if death is not None else ''
                    sex = sheet.cell(row=i, column=6).value
                    person = Person(name, s_name, l_name, birthday, death, sex)
                    if not person.valid():
                        continue
                    self.data.append(person)
                print(f"База данных загружена з файла {f_name}")
            except FileNotFoundError:
                print(f"Файл {f_name} відсутній")
            except ColumnException:
                print(f"У файлі {f_name} помилкова кількисть стовпчиків")
            except Exception:
                print(f"При читанні файла {f_name} пішло щось не так...")
        time.sleep(2)

    def __save(self, f_name):
        if f_name:
            wb = openpyxl.Workbook()
            sheet = wb.active
            for row_index, person in enumerate(self.data):
                values = (person.name, person.s_name, person.l_name,
                          person.birthday, person.death, person.sex)
                for col_index, value in enumerate(values):
                    cell = sheet.cell(row=row_index + 2, column=col_index + 1)
                    if row_index == 0:
                        cell_name = sheet.cell(row=1, column=col_index + 1)
                        cell_name.value = self.name_fields[col_index]
                    cell.value = value
            wb.save(f_name)
            print(f"База данных збережена у файл {f_name}")
            time.sleep(2)

    def __export(self, f_name):
        if f_name:
            json_data = []
            for person in self.data:
                json_dict = {}
                json_dict[self.name_fields[0]] = person.name
                json_dict[self.name_fields[1]] = person.s_name
                json_dict[self.name_fields[2]] = person.l_name
                json_dict[self.name_fields[3]] = person.birthday
                json_dict[self.name_fields[4]] = person.death
                json_dict[self.name_fields[5]] = person.sex
                json_data.append(json_dict)

            with open(f_name, 'w') as f:
                json.dump(json_data, f)
            print(f"База данных експортована у файл {f_name}")
            time.sleep(2)

    def input_data(self, *args):
        while True:
            name = input("Введите ім'я: ")
            s_name = input("Введите призвіще або ENTER: ")
            l_name = input("Введите по-батькові або ENTER: ")
            birthday = input("Введите дату народження (20.10.1980): ")
            death = input("Введите дату смерті (20.10.1980) або ENTER: ")
            sex = input("Введите стать (m/f): ")
            person = Person(name, s_name, l_name, birthday, death, sex)
            if not person.valid():
                print("Помилка ввода:", person.error)
                answer = input("Бажаєте повторити ввід (т/y): ")
                if answer.lower() not in ("y", "т"):
                    break
                continue
            self.data.append(person)
            print('-' * 50)
            answer = input("Бажаєте продовжувати ввід (т/y): ")
            if answer.lower() not in ("y", "т"):
                break

    def get_find(self, find_word):
        result = []
        for person in self.data:
            if find_word.lower() in person.full_name.lower():
                result.append(person)
        return result

    def find(self, *args):
        while True:
            find_word = input("Введите ім'я та/або призвіще та/або по-батькові: ")
            result = self.get_find(find_word)
            if result:
                for person in result:
                    print(person)
            else:
                print("По вашему запиту нічого не знайдено")
            print('-' * 50)
            answer = input("Бажаєте повторити пошук (т/y): ")
            if answer.lower() not in ("y", "т"):
                break

    def file(self, choice):
        f_name = input("Введите ім'я файлу: ")
        f_name = f"{f_name.split('.')[0]}{self.extensions[choice]}"
        self.choices[choice](f_name)
