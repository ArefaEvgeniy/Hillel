import openpyxl

wb = openpyxl.load_workbook('task_04+.xlsx')
sheet = wb.active
print(wb.sheetnames)
print(sheet)

print(sheet['C3'].value)

rows = sheet.max_row
cols = sheet.max_column

name_of_field = []
data = []
for i in range(1, rows+1):
    field = []
    for j in range(1, cols+1):
        cell = sheet.cell(row=i, column=j)
        if i == 1:
            name_of_field.append(cell.value)
        else:
            field.append(cell.value)
    if field:
        data.append(field)

print(name_of_field)
print(data)
