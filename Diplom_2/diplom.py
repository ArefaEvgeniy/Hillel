import tkinter as tk
from tkinter import ttk
from tkinter.messagebox import showinfo
import openpyxl
from openpyxl.styles import Font
from datetime import date


class Spinbox(ttk.Entry):

    def __init__(self, master=None, **kw):

        ttk.Entry.__init__(self, master, "ttk::spinbox", **kw)
    def set(self, value):
        self.tk.call(self._w, "set", value)


class MyExcept(Exception):
    pass


def update_days(*args):
    selected_month = date_month.get()
    if selected_month in month_list:
        days_in_month = month_list[selected_month]
        day_list = [str(i) for i in range(1, days_in_month + 1)]
        date_day.config(values=day_list)
        date_day.current(0)


def update_days_end_date(*args):
    selected_month = end_date_month.get()
    if selected_month in month_list:
        days_in_month = month_list[selected_month]
        end_day_list = [str(i) for i in range(1, days_in_month + 1)]
        end_date_day.config(values=end_day_list)
        end_date_day.current(0)


def create_new_file():
    work_book = openpyxl.Workbook()
    work_book.create_sheet('Пользователи', index=0)
    sheet = work_book['Пользователи']
    sheet.append(cols)

    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    work_book.save("Data.xlsx")


def save_append():
    try:
        workbook = openpyxl.load_workbook("Data.xlsx")
        sheet = workbook.active
        list_tree = list(treeview.get_children())
        print(list_tree)
        for item in list_tree:
            values = treeview.item(item, 'values')
            sheet.append(values)

        workbook.save("Data.xlsx")
    except FileNotFoundError:
        showinfo("Ошибка ввода", "Файл для записи не создан, создайте файл")


def save_overwrite():
    work_book = openpyxl.Workbook()
    work_book.create_sheet('Пользователи', index=0)
    sheet = work_book['Пользователи']
    sheet.append(cols)

    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    list_tree = list(treeview.get_children())
    print(list_tree)
    for item in list_tree:
        values = treeview.item(item, 'values')
        sheet.append(values)

    work_book.save("Data.xlsx")


def load_from_disk():
    try:
        workbook = openpyxl.load_workbook("Data.xlsx")
        sheet = workbook.active
        list_excel = list(sheet.values)

        for value_tuple in list_excel[1:]:
            treeview.insert('', tk.END, values=value_tuple)
    except FileNotFoundError:
        showinfo("Ошибка ввода", "Файл отсутствует, сначала сохраните файл")


def clear_list():
    treeview.delete(*treeview.get_children())
    start()


def start():
    for col_name in cols:
        treeview.heading(col_name, text=col_name)


def exceptions_date():
    val_end_date_year = end_date_year.get()
    list_date_month = end_date_month.get()
    for month in month_list_2.keys():
        if month == list_date_month:
            val_end_date_month = int(month_list_2[month])
    val_end_date_day = end_date_day.get()

    data_today = date.today()
    end_date = date(int(val_end_date_year), int(val_end_date_month),
                    int(val_end_date_day))
    day = data_today - end_date

    if not int(day.days) >= 0:
        # блок except находится в родительской функции
        raise MyExcept("Окончательная дата больше фактической даты календаря")


def exceptions_end_date():
    val_date_year = date_birth_year.get()
    list_date_birth_month = date_month.get()
    for month in month_list_2.keys():
        if month == list_date_birth_month:
            val_date_birth_month = month_list_2[month]
    val_date_day = date_day.get()
    val_end_date_year = end_date_year.get()
    list_date_month = end_date_month.get()
    for month in month_list_2.keys():
        if month == list_date_month:
            val_end_date_month = int(month_list_2[month])
    val_end_date_day = end_date_day.get()

    birth_date = date(int(val_date_year), int(val_date_birth_month),
                      int(val_date_day))
    end_date = date(int(val_end_date_year), int(val_end_date_month),
                    int(val_end_date_day))
    day = end_date - birth_date

    if not int(day.days) >= 0:
        # блок except находится в родительской функции
        raise MyExcept("Дата рождения не может быть больше окончательной даты")


def exceptions():
    val_first_name = first_name.get()
    val_first_name = "" if val_first_name == "Имя*" else val_first_name
    val_last_name = last_name.get()
    val_surname = surname.get()
    val_gender = gender.get()
    val_date_year = date_birth_year.get()
    val_date_birth_month = date_month.get()
    val_date_day = date_day.get()
    val_end_date_year = end_date_year.get()
    val_end_date_month = end_date_month.get()
    val_end_date_day = end_date_day.get()

    try:
        if val_first_name == "":
            raise MyExcept("Поле имя не может быть пустым")
        if not val_last_name or not val_surname:
            pass
        elif not val_first_name.isalpha() or not val_last_name.isalpha() \
                or not val_surname.isalpha():
            raise MyExcept("Поля ФИО не могут содержать символы или цифры")
        if val_gender == "None":
            raise MyExcept("Выберите пол")
        if not val_date_year.isdigit() or not 2023 >= int(val_date_year) >= 1950 \
                or not val_end_date_year.isdigit() or not 2023 >= int(val_end_date_year) >= 1951:
            raise MyExcept("Поле Год должен быть числом")
        if not val_date_birth_month in month_list.keys() or \
                not val_end_date_month in month_list.keys():
            raise MyExcept("Месяц выбран не верно")
        if not val_date_day.isdigit() or not 31 >= int(val_date_day) >= 0 or \
                not val_end_date_day.isdigit() or not 31 >= int(val_end_date_day) >= 0:
            raise MyExcept("День должен быть числом от 0 до 31")
        exceptions_date()  # проверка на фактическую дату
        exceptions_end_date()  # проверка даты рождения и окончательной даты
    except MyExcept as error:
        showinfo("Ошибка ввода", str(error))
        raise MyExcept


def birth_date():
    # приводим год, месяц, день в единую переменную
    val_date_birth_year = date_birth_year.get()
    list_date_birth_month = date_month.get()
    for month in month_list_2.keys():
        if month == list_date_birth_month:
            val_date_birth_month = str(month_list_2[month]).zfill(2)
    val_date_day = str(date_day.get()).zfill(2)

    val_date_birth = f"{val_date_day}.{val_date_birth_month}." \
                     f"{val_date_birth_year}"
    return val_date_birth


def end_date():
    # приводим год, месяц, день в единую переменную
    val_end_date_year = end_date_year.get()
    list_date_month = end_date_month.get()
    for month in month_list_2.keys():
        if month == list_date_month:
            val_end_date_month = str(month_list_2[month]).zfill(2)
    val_end_date_day = str(end_date_day.get()).zfill(2)

    val_end_date = f"{val_end_date_day}.{val_end_date_month}." \
                   f"{val_end_date_year}"
    return val_end_date


def calculate_age():
    val_date_birth_year = int(date_birth_year.get())
    list_date_birth_month = date_month.get()
    for month in month_list_2.keys():
        if month == list_date_birth_month:
            val_date_birth_month = int(month_list_2[month])
    val_date_day = int(date_day.get())

    val_end_date_year = int(end_date_year.get())
    list_date_month = end_date_month.get()
    for month in month_list_2.keys():
        if month == list_date_month:
            val_end_date_month = int(month_list_2[month])
    val_end_date_day = int(end_date_day.get())

    # формула расчета сколько полных лет
    day = val_end_date_day - val_date_day
    if day >= 0:
        month = val_end_date_month - val_date_birth_month
        if month >= 0:
            return val_end_date_year - val_date_birth_year
        elif month < 0:
            return val_end_date_year - 1 - val_date_birth_year

    elif day < 0:
        month = val_end_date_month - 1 - val_date_birth_month
        if month >= 0:
            return val_end_date_year - val_date_birth_year
        elif month < 0:
            return val_end_date_year - 1 - val_date_birth_year


def insert_row():
    try:
        exceptions()

        val_first_name = first_name.get()
        val_first_name = " " if val_first_name == "Имя*" else val_first_name
        val_last_name = last_name.get()
        val_last_name = " " if val_last_name == "Фамилия" else val_last_name
        val_surname = surname.get()
        val_surname = " " if val_surname == "Отчество" else val_surname
        val_gender = gender.get()
        # дата рождения
        val_date_birth = birth_date()
        # Дата окончания
        val_end_date = end_date()
        # Возраст
        age = calculate_age()

        print(val_first_name, val_last_name, val_surname, val_gender,
              val_date_birth, val_end_date, age)

        value_list = [val_first_name, val_last_name, val_surname, val_gender,
                      val_date_birth, val_end_date, age]
        treeview.insert('', tk.END, values=value_list)
        reset_input()

    except Exception:
        print("Ошибка с вводом обработана, всё работает корректно")


def reset_input():
    first_name.delete('0', 'end')
    first_name.insert(0, "Имя*")
    last_name.delete('0', 'end')
    last_name.insert(0, "Фамилия")
    surname.delete('0', 'end')
    surname.insert(0, "Отчество")
    gender.set(None)
    date_birth_year.delete('0', 'end')
    date_birth_year.insert(0, "2000")
    date_month.current(0)
    date_day.current(0)
    end_date_year.delete('0', 'end')
    end_date_year.insert(0, data_today.year)
    end_date_month.current(int(data_today.month - 1))
    end_date_day.current(int(data_today.day - 1))


def search_results():
    list_temp = []
    search_result = search_results_entry.get()
    list_tree = list(treeview.get_children())
    print(list_tree)
    for item in list_tree:
        value = treeview.item(item, 'values')
        for user in value:
            if search_result.lower() in user.lower():
                list_temp.append(value)

    list_temp = set(list_temp)
    clear_list()
    for item in list_temp:
        treeview.insert('', tk.END, values=item)


def delete_selected():
    selected_item = treeview.selection()
    if selected_item:
        treeview.delete(selected_item)


win = tk.Tk()

# Название и иконка
# иконка работает только у меня, по этому код ниже закомментирован
# photo = tk.PhotoImage(file='icon_hillel.png')
# win.iconphoto(False, photo)
win.config(bg="#D8F7F5")
win.title("База данных")
win.minsize(400, 350)

# тема
style = ttk.Style(win)
style.theme_use("alt")

#  меню Файл
menubar = tk.Menu(win)
win.config(menu=menubar)
setting_menu = tk.Menu(menubar, tearoff=0)
setting_menu.add_command(label="Создать новый файл", command=create_new_file)
setting_menu.add_command(label="Сохранить (дозапись)", command=save_append)
setting_menu.add_command(label="Сохранить (переписать)", command=save_overwrite)
setting_menu.add_command(label="Загрузить с диска", command=load_from_disk)
setting_menu.add_command(label="Выход", command=win.destroy)
menubar.add_cascade(label="Файл", menu=setting_menu)

# переменные
month_list = {"Январь": 31, "Февраль": 29, "Март": 31, "Апрель": 30, "Май": 31,
              "Июнь": 31, "Июль": 30, "Август": 31,
              "Сентябрь": 30, "Октябрь": 31, "Ноябрь": 30, "Декабрь": 31}

month_list_2 = {"Январь": 1, "Февраль": 2, "Март": 3, "Апрель": 4, "Май": 5,
                "Июнь": 6, "Июль": 7, "Август": 8, "Сентябрь": 9, "Октябрь": 10,
                "Ноябрь": 11, "Декабрь": 12}
cols = ("Имя", "Фамилия", "Отчество", "Пол", "Дата рождения", "Дата окончания",
        "Возраст")
data_today = date.today()

# Рамка с лева, Ввод пользователя
frame = ttk.Frame(win)
frame.grid(row=0, column=0, padx=5, pady=5)

widgets_frame = ttk.LabelFrame(frame, text="Ввод пользователя")
widgets_frame.grid(row=0, column=0, padx=10, pady=5, sticky="ew")

first_name = ttk.Entry(widgets_frame)
first_name.insert(0, "Имя*")
first_name.bind("<FocusIn>", lambda e: first_name.delete('0', 'end'))
first_name.grid(row=0, column=0, padx=5, pady=(0, 5), sticky="ew")

last_name = ttk.Entry(widgets_frame)
last_name.insert(0, "Фамилия")
last_name.bind("<FocusIn>", lambda e: last_name.delete('0', 'end'))
last_name.grid(row=1, column=0, padx=5, pady=(0, 5), sticky="ew")

surname = ttk.Entry(widgets_frame)
surname.insert(0, "Отчество")
surname.bind("<FocusIn>", lambda e: surname.delete('0', 'end'))
surname.grid(row=2, column=0, padx=5, pady=(0, 5), sticky="ew")

# Рамка выберите пол
frame_gender = ttk.LabelFrame(frame, text="Выберите пол")
frame_gender.grid(row=1, column=0, padx=10, pady=5)

# выберите пол
gender = tk.StringVar()
gender.set(None)
gender_male = tk.Radiobutton(frame_gender, text="Мужчина", variable=gender,
                             value="Мужчина")
gender_female = tk.Radiobutton(frame_gender, text="Женщина", variable=gender,
                               value="Женщина")
gender_male.grid(row=0, column=0)
gender_female.grid(row=0, column=1)

# Рамка дата рождения
frame_date_birth = ttk.LabelFrame(frame, text="Дата рождения")
frame_date_birth.grid(row=3, column=0, padx=10, pady=5)

# год рождения
date_birth_l = tk.Label(frame_date_birth, text="Год")
date_birth_l.grid(row=0, column=0, pady=(0, 5), sticky="ew")
date_birth_year = Spinbox(frame_date_birth, from_=1950, to=2023)
date_birth_year.insert(0, "2000")
date_birth_year.grid(row=0, column=1, padx=5, pady=(0, 5))

# месяц рождения
date_month_l = tk.Label(frame_date_birth, text="Месяц")
date_month_l.grid(row=1, column=0, pady=(0, 5))
date_month = ttk.Combobox(frame_date_birth, values=list(month_list.keys()))
date_month.current(0)
date_month.grid(row=1, column=1, padx=5, pady=(0, 5), sticky="ew")
date_month.bind("<<ComboboxSelected>>", update_days)

# день(число) рождения
date_day_l = tk.Label(frame_date_birth, text="День")
date_day_l.grid(row=2, column=0, pady=(0, 5), sticky="ew")
day_list = [str(i) for i in range(1, 32)]
date_day = ttk.Combobox(frame_date_birth, values=day_list)
date_day.current(0)
date_day.grid(row=2, column=1, padx=5, pady=(0, 5), sticky="ew")

# Рамка Дата окончания
frame_end_date = ttk.LabelFrame(frame, text="Дата окончания")
frame_end_date.grid(row=4, column=0, padx=10, pady=5)

# конечный год
end_date_l = tk.Label(frame_end_date, text="Год")
end_date_l.grid(row=0, column=0, pady=(0, 5), sticky="ew")
end_date_year = Spinbox(frame_end_date, from_=1950, to=2023)
end_date_year.insert(0, data_today.year)
end_date_year.grid(row=0, column=1, padx=5, pady=(0, 5))

# конечный месяц
end_date_month_l = tk.Label(frame_end_date, text="Месяц")
end_date_month_l.grid(row=1, column=0, pady=(0, 5))
end_date_month = ttk.Combobox(frame_end_date, values=list(month_list.keys()))
end_date_month.current(int(data_today.month - 1))
end_date_month.grid(row=1, column=1, padx=5, pady=(0, 5), sticky="ew")
end_date_month.bind("<<ComboboxSelected>>", update_days_end_date)

# конечный день
end_date_day_l = tk.Label(frame_end_date, text="День")
end_date_day_l.grid(row=2, column=0, pady=(0, 5), sticky="ew")
end_day_list = [str(i) for i in range(1, 32)]
end_date_day = ttk.Combobox(frame_end_date, values=day_list)
end_date_day.current(int(data_today.day - 1))
end_date_day.grid(row=2, column=1, padx=5, pady=(0, 5), sticky="ew")

# Кнопка сохранить пользователя
button = ttk.Button(frame, text="Сохранить пользователя", takefocus=False,
                    command=insert_row)
button.grid(row=5, column=0, padx=1, pady=(0, 5), sticky="ew")

# ПРАВАЯ ЧАСТЬ

# Рамка с права, список пользователей поиск, очистить поле и удалить пользователя
frame_on_the_right = ttk.Frame(win)
frame_on_the_right.grid(row=0, column=1, padx=5, pady=5)

# Рамка поиск
frame_search_results = ttk.Frame(frame_on_the_right)
frame_search_results.grid(row=0, column=0, padx=5, pady=5, sticky="we")

# текст поиск
search_results_label = tk.Label(frame_search_results, text="Поиск пользователя")
search_results_label.grid(row=0, column=0, padx=1, pady=(0, 5), sticky="we")

# поле ввода поиска
search_results_entry = ttk.Entry(frame_search_results, width=40)
search_results_entry.insert(0, "Введите имя либо фамилию либо отчество")
search_results_entry.bind("<FocusIn>", lambda e: search_results_entry.delete('0', 'end'))
search_results_entry.grid(row=0, column=1, padx=1, pady=(0, 5), sticky="we")

# кнопка поиска
search_results_button = ttk.Button(frame_search_results, text="Поиск",
                                   takefocus=False, command=search_results)
search_results_button.grid(row=0, column=2, padx=1, pady=(0, 5), sticky="we")

# кнопка очистить
clear_search_results_button = ttk.Button(frame_search_results, text="Очистить",
                                         takefocus=False, command=clear_list)
clear_search_results_button.grid(row=0, column=3, padx=1, pady=(0, 5), sticky="we")

# кнопка удалить
delete_button = ttk.Button(frame_search_results, text="Удалить",
                           takefocus=False, command=delete_selected)
delete_button.grid(row=0, column=4, padx=1, pady=(0, 5), sticky="we")

# Рамка с деревом пользователей
tree_frame = ttk.Frame(frame_on_the_right)
tree_frame.grid(row=1, column=0, padx=2, pady=1, sticky="sn")
tree_scroll = ttk.Scrollbar(tree_frame)  # создание полосы прокрутки
tree_scroll.pack(side="right", fill="y")

treeview = ttk.Treeview(tree_frame, show="headings",
                        yscrollcommand=tree_scroll.set, columns=cols, height=17)
treeview.column("Имя", width=80)
treeview.column("Фамилия", width=80)
treeview.column("Отчество", width=90)
treeview.column("Пол", width=80)
treeview.column("Дата рождения", width=100)
treeview.column("Дата окончания", width=100)
treeview.column("Возраст", width=70)
treeview.pack()
tree_scroll.config(command=treeview.yview)  # команда прокрутки

start()  # заполнение заглавных полей
win.mainloop()
